from openpyxl import load_workbook
from helpers.director.shortcut import director_view,get_request_cache
from django.conf import settings
import os
from helpers.director.shortcut import Fields,ModelFields

class ExcelFields(ModelFields):
    '''
    需要自定义  
    
    def save_form()
    '''
    
    def set_named_ctx(self):
        director_name = self.get_director_name()
        named_ctx = get_request_cache()['named_ctx']
        if director_name not in named_ctx:
            named_ctx.update({
                director_name:self.get_head_context()
            })
    
    def clean(self):
        pass
    
    def get_front_action(self,):
        """
        发送到前端执行的代码。
        `ctx.row.par_row_pk = scope.ps.vc.par_row.pk` 用于从前端环境提取变量信息，可以overwrite该行代码
        """
        self_director_name = self.get_director_name()
        return '''
        var ctx = named_ctx["%(director_name)s"]
        ex.uploadfile({accept:".xlsx"})
            .then((resp)=>{return ex.director("%(director_name)s").parse_excel_head({url:resp[0],par_row:scope.ps.vc.par_row.pk}) } )
            .then((resp)=>{ 
                ex.each(ctx.heads,head=>{head.options=resp.options});
                ctx.row=resp.row;
                ctx.row.par_row_pk = scope.ps.vc.par_row.pk
                cfg.pop_vue_com('com-form-one',ctx)
            }).then(()=>{
               scope.ps.search()
            })
        '''%{'director_name':self_director_name}
    
    def parse_excel_head(self,url,**kws):
        heads_names = get_excel_head(url)
        heads =[]
        for index,name in enumerate(heads_names):
            heads.append({'value':index,'label':name,})
        
        row ={
            '_director_name':self.get_director_name(),
            'meta__url':url
        }
        for head in self.get_heads():
            index = get_match_index(head['label'], heads_names)
            if index:
                row[head['name']] = index
        return {
            'options':heads,
            'row':row
            }
    
    def dict_head(self, head):
        head['editor']='com-field-select'
        head['group']='1'
        head['option_show'] = '''
        (function(){
            var heads = scope.ps.vc.heads
            for(var i=0;i<heads.length;i++){
                var head = heads[i]
               if( head.group != '1')continue
               var key = head.name
               if(key !=scope.vc.head.name && scope.row[key] ==scope.option.value){
                  return false
               }
             }
           return true
        })()
        '''
        return head
    
    def parser_excel(self,url,dispatch_head,valid_row=[]):
        '''
        @valid_row: list  每行，必须要有该字段才算是有效行 
        '''
        return parser_excel(url, dispatch_head,valid_row)
    #def save_form(self):
        #url = self.kw.pop('meta__url')
        #record = self.kw.pop('record')
        #out_list = []
        #for row in parser_excel(url,self.kw,valid_row=['id_code']):
            #row.update({
                 #'record_id':record
            #} )
            #out_list.append(Salary(**row))
        ##print(out_list)
        #try:
            #Salary.objects.bulk_create(out_list)
        #except Exception as e:
            #raise UserWarning(str(e))

@director_view('func.get_excel_head')
def get_excel_head(url):
    path = url2path(url)
    wb = load_workbook(filename=path)
    
    sheets = wb.get_sheet_names()   # 获取所有表格(worksheet)的名字
    sheet0 = sheets[0]  # 第一个表格的名称
    ws = wb.get_sheet_by_name(sheet0) # 获取特定的 worksheet
    heads_names = [x.value.strip() if x.value else '' for x in next(ws.rows) ]
    
    #heads = [x.value for x in next(ws.rows)]
    
    return heads_names

def url2path(url):
    path = os.path.join( os.path.dirname(settings.BASE_DIR),url[1:] )
    return path

def get_match_index(head_name,heads):
    if head_name in heads:
        return heads.index(head_name)
    #for head in form_heads:
    return None

def parser_excel(url,dispatch_head,valid_row=[]):
    path = url2path(url)
    wb = load_workbook(filename=path,data_only=True)
    
    sheets = wb.get_sheet_names()   # 获取所有表格(worksheet)的名字
    sheet0 = sheets[0]  # 第一个表格的名称
    ws = wb.get_sheet_by_name(sheet0) # 获取特定的 worksheet
    count =0
    for row in ws.rows:
        if count >0:
            dc={}
            for k,v in dispatch_head.items():
                if not k.startswith(('_','meta_')):
                    index = int(v)
                    
                    dc[k]=row[index].value
                    #if row[index].value is not None:
                        #dc[k]=row[index].value
            
            is_valid = True
            for ff in valid_row:
                if dc.get(ff,None) is None:
                    is_valid=False
                    break
            if is_valid:
                yield dc
            else:
                continue
        else:
            count =1